import os
import re
import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
import logging

try:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import google.generativeai as genai
    GENAI_AVAILABLE = True
except ImportError:
    GENAI_AVAILABLE = False

# Setup basic logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DriverManager:
    """Manages Selenium WebDriver setup and teardown."""
    
    def __init__(self, headless=False, timeout=30):
        self.headless = headless
        self.timeout = timeout
        self.driver = None

    def get_driver(self):
        if self.driver is None:
            options = Options()
            if self.headless:
                options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.page_load_strategy = 'eager'
            
            try:
                self.driver = webdriver.Chrome(options=options)
                self.driver.set_page_load_timeout(self.timeout)
                logger.info("WebDriver initialized successfully.")
            except WebDriverException as e:
                logger.error(f"Failed to initialize WebDriver: {e}")
                raise
        return self.driver

    def close_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None
            logger.info("WebDriver closed.")

            

class SessionManager:
    """Manages requests.Session with retries and headers."""
    
    def __init__(self):
        self.session = self._setup_session()

    def _setup_session(self):
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1"
        })
        
        retry_strategy = Retry(
            total=5,
            backoff_factor=2,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        
        return session

    def get_session(self):
        return self.session

class DescriptionSummarizer:
    """Summarizes product descriptions using LLM (Gemini API)."""
    
    def __init__(self, api_key=None):
        self.api_key = api_key or os.environ.get("GEMINI_API_KEY")
        self.enabled = False
        
        if self.api_key and GENAI_AVAILABLE:
            try:
                genai.configure(api_key=self.api_key)
                # Using gemini-1.5-flash for fast and cost-effective text tasks
                self.model = genai.GenerativeModel('gemini-1.5-flash')
                self.enabled = True
                logger.info("DescriptionSummarizer initialized successfully with Gemini API.")
            except Exception as e:
                logger.warning(f"Failed to configure Gemini API: {e}. Summarization will be disabled.")
        else:
            if not GENAI_AVAILABLE:
                logger.warning("google-generativeai is not installed. Summarization is disabled. Run: pip install google-generativeai")
            elif not self.api_key:
                logger.warning("GEMINI_API_KEY not found in environment. Summarization is disabled.")

    def summarize(self, description):
        """Returns summarized text if enabled, else returns None."""
        if not self.enabled or not description or description == "N/A":
            return None
            
        # If description is already very short, no need to summarize
        if len(description) < 200:
            return None
            
        prompt = (
            "You are a professional medical product description summarizer. "
            "Please provide a very short, clean, and professional summary (1-3 sentences) of the following product description. "
            "Do not include any introductory or concluding remarks, just the summary.\n\n"
            f"Description:\n{description}"
        )
        
        try:
            response = self.model.generate_content(prompt)
            if response.text:
                return response.text.strip()
        except Exception as e:
            logger.error(f"Summarization failed: {e}")
            
        return None

class ProductLinkFetcher:
    """Navigates search pages and fetches product links."""
    
    def __init__(self, driver_manager):
        self.driver_manager = driver_manager

    def fetch(self, base_url, max_products):
        driver = self.driver_manager.get_driver()
        links = []
        logger.info(f"Loading search page: {base_url}")
        
        try:
            driver.get(base_url)
        except TimeoutException:
            logger.warning("Page load timed out, attempting to proceed...")
            
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/otc/'], a[href*='/drugs/']"))
            )
        except TimeoutException:
            logger.warning("Timeout waiting for initial product links to load.")
            
        time.sleep(3)
        
        if "Just a moment..." in driver.page_source or "Cloudflare" in driver.page_source:
            logger.error("Cloudflare bot protection triggered on search page.")
            return links
            
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts = 50
        
        while len(links) < max_products and scroll_attempts < max_scroll_attempts:
            soup = BeautifulSoup(driver.page_source, "html.parser")
            a_tags = soup.find_all("a", href=True)
            new_links_found = False
            
            for tag in a_tags:
                href = tag['href']
                if "/otc/" in href or "/drugs/" in href:
                    full_url = href if href.startswith("http") else f"https://www.1mg.com{href}"
                    if full_url not in links:
                        links.append(full_url)
                        new_links_found = True
                        if len(links) >= max_products:
                            break
                            
            logger.info(f"Collected {len(links)} links so far...")
            
            if len(links) >= max_products:
                break
                
            product_elements = driver.find_elements(By.CSS_SELECTOR, "a[href*='/otc/'], a[href*='/drugs/']")
            if product_elements:
                last_product = product_elements[-1]
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", last_product)
            else:
                driver.execute_script("window.scrollBy(0, 500);")
                
            time.sleep(3)
            new_height = driver.execute_script("return document.body.scrollHeight")
            
            if new_height == last_height and not new_links_found:
                driver.execute_script("window.scrollBy(0, 200);")
                time.sleep(3)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    logger.info("Reached end of page or no more products loading.")
                    break
                    
            last_height = new_height
            scroll_attempts += 1
                
        return links[:max_products]

class ProductParser:
    """Extracts detailed information from a single product page."""
    
    def __init__(self, driver_manager):
        self.driver_manager = driver_manager

    def parse(self, url):
        driver = self.driver_manager.get_driver()
        
        # Add retry loop for page load to handle occasional failures gracefully
        max_retries = 2
        for attempt in range(max_retries):
            try:
                driver.get(url)
                # Wait explicitly for either h1 or description to ensure main content has loaded
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.TAG_NAME, "h1"))
                )
                break
            except TimeoutException:
                logger.warning(f"Timeout on {url} (Attempt {attempt+1}/{max_retries})")
                if attempt == max_retries - 1:
                    logger.warning(f"Proceeding anyway for {url}")
            except Exception as e:
                logger.error(f"Error loading {url}: {e}")
                return None
                
        # Additional explicit wait for description area which might take longer
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div#aboutexpand, div.ProductDescription, div[class*='description']"))
            )
        except TimeoutException:
            # If description doesn't appear, we still proceed to parse whatever is there
            pass

        time.sleep(1)
        
        if "Just a moment..." in driver.page_source or "Cloudflare" in driver.page_source:
            logger.error("Cloudflare bot protection triggered on product page.")
            return None

        soup = BeautifulSoup(driver.page_source, "html.parser")
        
        company_name = "Mankind"
        medicine_name = "N/A"
        composition = ""
        price = "N/A"
        description_text = "N/A"
        image_urls = set()
        
        # 1. Medicine Name
        h1 = soup.find("h1")
        if h1:
            medicine_name = h1.get_text(strip=True)
        elif soup.title:
            medicine_name = soup.title.get_text(strip=True).split('|')[0].strip()
            
        # 2. Composition
        salt_tags = soup.find_all("a", href=lambda h: h and "/generics/" in h)
        if salt_tags:
            salts = []
            for tag in salt_tags:
                txt = tag.get_text(strip=True)
                if txt and txt not in salts:
                    salts.append(txt)
            composition = " + ".join(salts)
        else:
            salt_div = soup.find(lambda tag: tag.name == "div" and tag.get("class") and any("salt" in c.lower() for c in tag.get("class", [])))
            if salt_div:
                composition = salt_div.get_text(strip=True)
            
        # 3. Price
        price_text_found = ""
        price_divs = soup.find_all(lambda tag: tag.name in ["span", "div"] and tag.get("class") and any("price" in c.lower() or "mrp" in c.lower() for c in tag.get("class", [])))
        for div in price_divs:
            t = div.get_text(separator=' ', strip=True)
            if '₹' in t:
                price_text_found = t
                break
                
        if not price_text_found:
            for tag in soup.find_all(["span", "div"]):
                t = tag.get_text(separator=' ', strip=True)
                if '₹' in t and len(t) < 40 and ('mrp' in t.lower() or 'price' in t.lower()):
                    price_text_found = t
                    break
                    
        if price_text_found:
            match = re.search(r'₹\s?([0-9,.]+)', price_text_found)
            if match:
                price_str = match.group(1).replace(',', '').strip()
                try: 
                    price = float(price_str)
                except ValueError: 
                    price = price_str
                     
        # 4. Description
        desc_div = soup.find("div", {"id": "aboutexpand"})
        if not desc_div:
            desc_div = soup.find("div", {"class": re.compile(r"ProductDescription|product-description", re.I)})
        if not desc_div:
            desc_div = soup.find(lambda tag: tag.name == "div" and tag.get("class") and any("description" in c.lower() for c in tag.get("class", [])))
            
        if desc_div:
             raw_text = " ".join(desc_div.stripped_strings)
             clean_text = re.sub(r'\s+', ' ', raw_text).strip()
             if clean_text:
                 description_text = f"{medicine_name}\n{clean_text}"
             
        # 5. Images
        img_tags = soup.find_all("img")
        for img in img_tags:
            src = img.get("src")
            if not src or "http" not in src:
                continue
            
            classes = img.get("class", [])
            is_product_img = any("picture-image" in c.lower() or "thumbnail" in c.lower() for c in classes)
            
            if ("/image/upload/" in src) or is_product_img or ("w_380" in src) or ("w_700" in src):
                if not any(icon in src.lower() for icon in ['facebook', 'twitter', 'linkedin', 'instagram', 'youtube']):
                    image_urls.add(src)
            
        return {
            "companyName": company_name,
            "medicineName": medicine_name,
            "composition": composition,
            "price": price,
            "description": description_text,
            "image_urls": list(image_urls)
        }

class ImageDownloader:
    """Downloads images for products to the local filesystem."""
    
    def __init__(self, session_manager):
        self.session_manager = session_manager

    def clean_filename(self, name):
        return re.sub(r'[^a-zA-Z0-9]+', '', name)

    def download(self, image_urls, product_name, save_dir):
        if not image_urls:
            return ""
            
        cleaned_name = self.clean_filename(product_name)
        if not cleaned_name:
            cleaned_name = "product"
            
        os.makedirs(save_dir, exist_ok=True)
        saved_filenames = []
        image_urls = list(image_urls)[:10]
        session = self.session_manager.get_session()
        
        for i, url in enumerate(image_urls):
            filename = f"{cleaned_name}.png" if i == 0 else f"{cleaned_name}{i}.png"
            filepath = os.path.join(save_dir, filename)
            
            if filename in saved_filenames or os.path.exists(filepath):
                 saved_filenames.append(filename)
                 continue
                
            try:
                r = session.get(url, stream=True, timeout=10)
                r.raise_for_status()
                with open(filepath, 'wb') as f:
                    for chunk in r.iter_content(1024):
                        f.write(chunk)
                saved_filenames.append(filename)
                time.sleep(0.2)
            except Exception as e:
                logger.error(f"Failed to download image {url}: {e}")
                
        return ",".join(saved_filenames)

class ExcelManager:
    """Handles formatting and saving data to Excel."""
    
    def __init__(self):
        pass

    def save(self, data, output_file):
        columns_order = ["companyName", "medicineName", "composition", "price", "description", "imageLink"]
        df = pd.DataFrame(data, columns=columns_order)
        
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        if not OPENPYXL_AVAILABLE:
            logger.error("Please install openpyxl to format and save as Excel. Falling back to CSV.")
            df.to_csv(output_file.replace('.xlsx', '.csv'), index=False)
            return
            
        try:
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Products')
            
            workbook = writer.book
            worksheet = writer.sheets['Products']
            
            # Apply formatting
            for i, col_name in enumerate(df.columns, start=1):
                column_letter = get_column_letter(i)
                
<<<<<<< HEAD
            img_urls = product_info.pop("image_urls", [])
            # still use requests session for fast image downloading
            saved_filenames_str = download_images(session, img_urls, product_info['medicineName'], images_dir)
            
            if saved_filenames_str:
                formatted_paths = [f"./images/{fname}" for fname in saved_filenames_str.split(',')]
                product_info['imageLink'] = ",\n".join(formatted_paths)
            else:
                product_info['imageLink'] = ""
=======
                # Auto width sizing logic. Description column is usually wide, limit to 60.
                if col_name == "description":
                    worksheet.column_dimensions[column_letter].width = 60
                else:
                    max_len = max(df[col_name].astype(str).map(len).max(), len(col_name))
                    worksheet.column_dimensions[column_letter].width = min(max_len + 2, 40)
                
            # Wrap text and set top alignment for all cells
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
            writer.close()
            logger.info(f"Data saved successfully to {output_file}")
>>>>>>> dce1a23 (The current `scraper.py` already contains the requested classes, but they are all clumped together in one file. To truly achieve the requested goal ("Separate each class into different sections clearly so I can easily access, edit, and maintain them"), I propose **splitting the classes into separate Python files (modules))
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")

class MainScraper:
    """Controller class orchestrating the entire scraping process."""
    
    def __init__(self):
        load_dotenv()
        self.base_url = os.environ.get("BASE_URL", "https://www.1mg.com/search/all?name=mankind")
        self.max_products = int(os.environ.get("MAX_PRODUCTS", "100"))
        self.output_dir = os.environ.get("OUTPUT_FOLDER", "mankind")
        self.images_dir = os.path.join(self.output_dir, "images")
        self.output_excel = os.path.join(self.output_dir, "data.xlsx")
        self.is_headless = os.environ.get("HEADLESS", "False").lower() == "true"
        
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)
        
        self.driver_manager = DriverManager(headless=self.is_headless)
        self.session_manager = SessionManager()
        
        self.link_fetcher = ProductLinkFetcher(self.driver_manager)
        self.parser = ProductParser(self.driver_manager)
        self.downloader = ImageDownloader(self.session_manager)
        self.summarizer = DescriptionSummarizer()
        self.excel_manager = ExcelManager()

    def run(self):
        try:
            logger.info(f"Fetching up to {self.max_products} product links...")
            product_links = self.link_fetcher.fetch(self.base_url, self.max_products)
            logger.info(f"Gathered {len(product_links)} product links.")
            
            if not product_links:
                logger.warning("No products to process.")
                return
                
            all_products_data = []
            
            for i, link in enumerate(product_links, 1):
                logger.info(f"[{i}/{len(product_links)}] Scraping: {link}")
                product_info = self.parser.parse(link)
                
                if not product_info:
                    logger.warning("Failed to extract data, skipping.")
                    continue
                    
                # Handle Image Downloading
                img_urls = product_info.pop("image_urls", [])
                self.downloader.download(img_urls, product_info['medicineName'], self.images_dir)
                product_info['imageLink'] = ",\n".join(img_urls) if img_urls else ""
                
                # Handle LLM Summarization and formatting
                original_desc = product_info['description']
                summary = self.summarizer.summarize(original_desc)
                
                if summary:
                    formatted_desc = f"{summary}\n\nRead More...\n\n{original_desc}"
                    product_info['description'] = formatted_desc
                
                all_products_data.append(product_info)
                
            if all_products_data:
                self.excel_manager.save(all_products_data, self.output_excel)
            else:
                logger.warning("No product data collected.")
                
        except Exception as e:
            logger.error(f"An unexpected error occurred during execution: {e}", exc_info=True)
        finally:
            if not self.is_headless:
                logger.info("Browser is still open for you to inspect.")
                input("Press Enter in this terminal to close browser...")
            self.driver_manager.close_driver()

if __name__ == "__main__":
    scraper = MainScraper()
    scraper.run()
