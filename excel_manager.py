import os
import pandas as pd
from utils import setup_logger

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = setup_logger(__name__)

class ExcelManager:
    """Handles formatting and saving data to Excel."""
    
    def __init__(self):
        pass

    def save(self, data, output_file):
        columns_order = ["companyName", "medicineName", "composition", "price", "description", "imageLink"]
        df = pd.DataFrame(data, columns=columns_order)
        
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        if not OPENPYXL_AVAILABLE:
            logger.error("Please install openpyxl to format and save as Excel. Falling back to CSV.")
            df.to_csv(output_file.replace('.xlsx', '.csv'), index=False)
            return
            
        try:
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Products')
            
            workbook = writer.book
            worksheet = writer.sheets['Products']
            
            # Define styles
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True, size=12)
            
            # Format headers
            for col_num in range(1, len(columns_order) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply formatting for columns
            for i, col_name in enumerate(df.columns, start=1):
                column_letter = get_column_letter(i)
                
                # Auto width sizing logic.
                if col_name == "description":
                    worksheet.column_dimensions[column_letter].width = 80 # Much wider for long text and summaries
                elif col_name == "imageLink":
                    worksheet.column_dimensions[column_letter].width = 30 # Wider for multiple image names
                else:
                    max_len = max(df[col_name].astype(str).map(len).max(), len(col_name))
                    worksheet.column_dimensions[column_letter].width = min(max_len + 5, 50)
                
            # Format data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border
                    
            writer.close()
            logger.info(f"Data saved successfully to {output_file} with professional formatting.")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
